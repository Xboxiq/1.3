#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""مزامنة تدريجية من أرشيف نماذج خدمات المشتركين (RAR/ZIP).

- يستخرج جدول أسعار 2026 ويربطه بكل خدمة (CS/CB/CT/CA)
- يحدّث حقول pricing و guide في data/services.json فقط
- ينسخ مخططات الانسيابي المحدّثة إلى data/flowcharts/
- لا يستبدل قوالب Word (تحافظ على علامات {{...}})

التشغيل:
    python3 tools/sync_services_from_archive.py /path/to/extracted/archive
    python3 tools/sync_services_from_archive.py   # يفكّ RAR الافتراضي في الجذر
"""
from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SERVICES_JSON = ROOT / 'data' / 'services.json'
FLOWCHARTS = ROOT / 'data' / 'flowcharts'
SOURCE_DIR = ROOT / 'data' / 'source'

# ربط رمز الخدمة بجدول الأسعار 2026 (priceId من Excel)
SERVICE_PRICE_MAP: dict[str, dict] = {
    'CS0001': {'mode': 'by_class', 'priceIds': {'منزلي': 3, 'تجاري': 4, 'زراعي': 5, 'صناعي': 6, 'حكومي': 7, 'مجمع سكني': 3, 'مشروع استثماري': 4}, 'variablePriceId': 8, 'label': 'أجور الكشف + كلفة الإيصال'},
    'CS0002': {'mode': 'fixed', 'priceId': 55, 'label': 'رسوم طلب الخدمة'},
    'CS0003': {'mode': 'by_class', 'priceIds': {'منزلي': 3, 'تجاري': 4, 'زراعي': 5, 'صناعي': 6, 'حكومي': 7, 'مجمع سكني': 3, 'مشروع استثماري': 4}, 'label': 'أجور الكشف (حسب الصنف الجديد)'},
    'CS0004': {'mode': 'by_class_phase', 'priceIds': {'منزلي': (14, 15), 'تجاري': (22, 23), 'زراعي': (20, 21), 'صناعي': (18, 19), 'حكومي': (16, 17)}, 'label': 'أجور القطع والإعادة'},
    'CS0005': {'mode': 'by_class_phase', 'priceIds': {'منزلي': (14, 15), 'تجاري': (22, 23), 'زراعي': (20, 21), 'صناعي': (18, 19), 'حكومي': (16, 17)}, 'label': 'أجور إعادة التيار'},
    'CS0006': {'mode': 'fixed', 'priceId': 12, 'label': 'رسوم إلغاء الاشتراك / الهدم'},
    'CS0007': {'mode': 'none', 'label': 'بدون رسوم ثابتة — مراجعة إدارية'},
    'CS0008': {'mode': 'fixed', 'priceId': 55, 'label': 'رسوم تعديل بيانات المشترك'},
    'CS0009': {'mode': 'fixed', 'priceId': 36, 'label': 'قراءة مقياس بناءً على طلب المشترك'},
    'CS0010': {'mode': 'compound', 'priceIds': [58, 59, 60], 'label': 'إيصال قدرة مؤقتة + أجور يومية'},
    'CS0011': {'mode': 'fixed', 'priceId': 55, 'label': 'رسوم نقل ملكية / تعديل بيانات'},
    'CB0001': {'mode': 'none', 'label': 'دفع قائمة الأجور — بدون رسوم خدمة إضافية'},
    'CB0002': {'mode': 'none', 'label': 'تقرير استهلاك — يُحدد حسب الدراسة'},
    'CB0003': {'mode': 'none', 'label': 'مراجعة قائمة — بدون رسوم ثابتة'},
    'CB0004': {'mode': 'by_class', 'priceIds': {'منزلي': 44, 'تجاري': 45, 'صناعي': 45, 'زراعي': 45, 'حكومي': 45, 'مجمع سكني': 45, 'مشروع استثماري': 45}, 'label': 'بدل فاقد / جباية موقفة'},
    'CB0005': {'mode': 'by_class', 'priceIds': {'منزلي': 44, 'تجاري': 45, 'صناعي': 45, 'زراعي': 45, 'حكومي': 45, 'مجمع سكني': 45, 'مشروع استثماري': 45}, 'label': 'جباية موقفة'},
    'CB0006': {'mode': 'none', 'label': 'تسوية مالية — بدون رسوم ثابتة'},
    'CT0001': {'mode': 'fixed', 'priceId': 9, 'label': 'تغيير موقع العمود'},
    'CT0002': {'mode': 'fixed', 'priceId': 10, 'label': 'تغيير موقع ركيزة المحولة'},
    'CT0003': {'mode': 'variable', 'priceId': 8, 'label': 'كلفة الإيصال / تغيير الكابل'},
    'CT0004': {'mode': 'variable', 'priceId': 8, 'label': 'كلفة تجزئة الاشتراك'},
    'CT0005': {'mode': 'variable', 'priceId': 8, 'label': 'كلفة تجميع الأحمال'},
    'CT0006': {'mode': 'fixed', 'priceId': 24, 'label': 'قطع وإعادة جهد 11 ك.ف.'},
    'CT0007': {'mode': 'meter_replace', 'label': 'استبدال / تغيير نوع المقياس'},
    'CT0008': {'mode': 'fixed', 'priceId': 12, 'label': 'رفع المقياس أو تغيير موقعه'},
    'CT0009': {'mode': 'meter_inspection', 'label': 'فحص / تبديل / صيانة المقياس'},
    'CT0010': {'mode': 'fixed', 'priceId': 12, 'label': 'نقل موقع الاشتراك داخل العقار'},
    'CA0001': {'mode': 'tamper', 'label': 'غرامات التلاعب (حسب الصنف)'},
    'CA0002': {'mode': 'none', 'label': 'إبلاغ خطر — بدون رسوم ثابتة'},
    'CA0003': {'mode': 'variable', 'label': 'تقدير أضرار الشبكة — بعد الكشف'},
    'CA0004': {'mode': 'none', 'label': 'شكوى إدارية — بدون رسوم ثابتة'},
}

METER_REPLACE_IDS = {
    ('منزلي', 'احادي'): 25, ('منزلي', 'ثلاثي'): 26,
    ('صناعي', 'احادي'): 27, ('صناعي', 'ثلاثي_مباشر'): 28, ('صناعي', 'ثلاثي_غير_مباشر'): 29,
    ('زراعي', 'احادي'): 30, ('زراعي', 'ثلاثي'): 31,
    ('حكومي', 'احادي'): 32, ('حكومي', 'ثلاثي'): 33,
    ('تجاري', 'احادي'): 34, ('تجاري', 'ثلاثي'): 35,
}
METER_INSPECTION_IDS = {('مقر', 'احادي'): 45, ('موقع', 'احادي'): 46, ('مقر', 'ثلاثي'): 47, ('موقع', 'ثلاثي'): 48}


def extract_rar(archive: Path, dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)
    for unrar in ('/tmp/unrar/unrar', 'unrar', 'unar'):
        try:
            subprocess.run(
                [unrar, 'x', '-o+', str(archive), str(dest) + '/'],
                check=True, capture_output=True,
            )
            return
        except (FileNotFoundError, subprocess.CalledProcessError):
            continue
    js = '''
const { createExtractorFromFile } = require("node-unrar-js");
(async () => {
  const ext = await createExtractorFromFile({ filepath: process.argv[1], targetPath: process.argv[2] });
  [...ext.extract().files];
})();
'''
    tmp = ROOT / 'tools' / '_unrar_once.js'
    tmp.write_text(js, encoding='utf-8')
    subprocess.run(['node', str(tmp), str(archive), str(dest)], check=True, cwd=ROOT)


def find_archive_root(base: Path) -> Path:
    for p in base.rglob('جدول اسعار الخدمات لسنة -2026 new.xlsx'):
        return p.parent
    raise FileNotFoundError('لم يُعثر على جذر الأرشيف (جدول الأسعار)')


def load_price_catalog(xlsx: Path) -> dict[int, dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    catalog: dict[int, dict] = {}
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        if seq is None:
            continue
        pid = int(ws.cell(r, 3).value) if ws.cell(r, 3).value is not None else int(seq)
        catalog[pid] = {
            'id': pid,
            'seq': int(seq),
            'category': str(ws.cell(r, 2).value or '').strip(),
            'service': str(ws.cell(r, 4).value or '').strip(),
            'amount2023': ws.cell(r, 5).value,
            'amount2026': ws.cell(r, 6).value,
            'notes': str(ws.cell(r, 7).value or '').strip() or None,
        }
    return catalog


def norm_code(raw: str) -> str | None:
    if not raw:
        return None
    m = re.search(r'([A-Z]{2})\s*0*(\d{4})', str(raw).upper())
    return f'{m.group(1)}{m.group(2)}' if m else None


def build_pricing(code: str, spec: dict, catalog: dict[int, dict]) -> dict:
    year = 2026
    base = {'year': year, 'currency': 'IQD', 'label': spec.get('label', 'رسوم الخدمة')}
    mode = spec.get('mode', 'none')

    def row(pid: int) -> dict:
        item = catalog.get(pid, {})
        amt = item.get('amount2026')
        return {
            'priceId': pid,
            'service': item.get('service'),
            'amount': amt if isinstance(amt, (int, float)) else None,
            'amountText': str(amt) if amt is not None and not isinstance(amt, (int, float)) else None,
            'notes': item.get('notes'),
        }

    if mode == 'none':
        return {**base, 'mode': 'none', 'display': 'بدون رسوم ثابتة'}
    if mode == 'fixed':
        r = row(spec['priceId'])
        return {**base, 'mode': 'fixed', **r, 'display': format_amount(r.get('amount'), r.get('amountText'))}
    if mode == 'variable':
        pid = spec.get('priceId')
        r = row(pid) if pid else {'amount': None, 'amountText': None, 'notes': None}
        return {**base, 'mode': 'variable', **r, 'display': r.get('amountText') or 'تُحدد بعد الدراسة الفنية'}
    if mode == 'by_class':
        tiers = {k: row(v) for k, v in spec['priceIds'].items()}
        return {**base, 'mode': 'by_class', 'tiers': tiers, 'display': 'حسب صنف الاشتراك'}
    if mode == 'by_class_phase':
        tiers = {}
        for cls, pair in spec['priceIds'].items():
            tiers[cls] = {'single': row(pair[0]), 'three': row(pair[1])}
        return {**base, 'mode': 'by_class_phase', 'tiers': tiers, 'display': 'حسب الصنف ونوع الربط'}
    if mode == 'compound':
        parts = [row(pid) for pid in spec['priceIds']]
        return {**base, 'mode': 'compound', 'parts': parts, 'display': ' + '.join(format_amount(p.get('amount'), p.get('amountText')) for p in parts)}
    if mode == 'meter_replace':
        tiers = {f'{k[0]}_{k[1]}': row(v) for k, v in METER_REPLACE_IDS.items()}
        return {**base, 'mode': 'meter_replace', 'tiers': tiers, 'display': 'حسب الصنف ونوع المقياس'}
    if mode == 'meter_inspection':
        tiers = {f'{k[0]}_{k[1]}': row(v) for k, v in METER_INSPECTION_IDS.items()}
        return {**base, 'mode': 'meter_inspection', 'tiers': tiers, 'display': 'حسب موقع الفحص ونوع المقياس'}
    if mode == 'tamper':
        tiers = {
            'منزلي_تجاوز': row(37), 'تجاري_تجاوز': row(38), 'تجاري_تجاوز_كبير': row(39),
            'منزلي_تلاعب': row(40), 'تجاري_تلاعب': row(41), 'تجاري_تلاعب_كبير': row(42),
        }
        return {**base, 'mode': 'tamper', 'tiers': tiers, 'display': 'حسب نوع المخالفة والصنف'}
    return {**base, 'mode': 'none', 'display': '—'}


def format_amount(amt, text=None) -> str:
    if isinstance(amt, (int, float)):
        return f'{int(amt):,}'.replace(',', '٬') + ' د.ع'
    return text or 'تُحدد لاحقاً'


def extract_guide_from_docx(path: Path) -> dict | None:
    sys.path.insert(0, str(ROOT / 'tools'))
    from docx_extract import extract
    raw = extract(str(path))
    proc_text = ''
    title = ''
    for line in raw.splitlines():
        if '|' not in line or line.startswith('==='):
            continue
        parts = [p.strip() for p in line.split('|') if p.strip()]
        if len(parts) >= 2 and not parts[0].isdigit():
            title = parts[0]
            proc_text = parts[-1]
            break
    if not proc_text:
        return None
    steps = [s.strip() for s in re.split(r'(?<=[.۔])\s+', proc_text) if len(s.strip()) > 8]
    return {'title': title, 'procedure': steps}


def find_procedural_doc(code: str, root: Path) -> Path | None:
    pat = re.compile(rf'{code}', re.I)
    cands = []
    for p in root.rglob('*.docx'):
        if p.name.startswith('~$'):
            continue
        name = p.name.lower()
        if 'وصف' in name or 'اجراء' in name or 'اجرائي' in name or 'وظيفي' in name:
            if pat.search(p.as_posix()):
                cands.append(p)
    return cands[0] if cands else None


def find_flowchart(code: str, root: Path) -> Path | None:
    exts = ('.png', '.jpg', '.jpeg', '.gif', '.webp')
    pat = re.compile(rf'{code}', re.I)
    for p in root.rglob('*'):
        if p.suffix.lower() in exts and pat.search(p.name):
            return p
    return None


def sync_flowcharts(root: Path, services: dict) -> int:
    FLOWCHARTS.mkdir(parents=True, exist_ok=True)
    n = 0
    for code in services:
        src = find_flowchart(code, root)
        if not src:
            continue
        dest = FLOWCHARTS / f'{code}{src.suffix.lower()}'
        if dest.suffix == '.jpeg':
            dest = FLOWCHARTS / f'{code}.jpg'
        shutil.copy2(src, dest)
        g = services[code].setdefault('guide', {})
        if isinstance(g.get('flowchart'), list):
            if g['flowchart']:
                g['flowchart'][0]['src'] = f'data/flowcharts/{dest.name}'
        else:
            g['flowchart'] = f'data/flowcharts/{dest.name}'
        n += 1
    return n


def main():
    archive_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    work = Path('/tmp/rar-forms-sync')
    if archive_arg and archive_arg.is_dir():
        root = find_archive_root(archive_arg)
    else:
        archive = archive_arg or next(ROOT.glob('*-1.rar'), None) or next(ROOT.glob('*.rar'), None)
        if not archive:
            archive = ROOT / 'نماذج طلبات  خدمات المشتركين.zip'
        if archive.suffix.lower() == '.rar':
            if work.exists():
                shutil.rmtree(work)
            extract_rar(archive, work)
            root = find_archive_root(work)
        elif archive.suffix.lower() == '.zip':
            if work.exists():
                shutil.rmtree(work)
            work.mkdir(parents=True)
            with zipfile.ZipFile(archive) as z:
                z.extractall(work)
            root = find_archive_root(work)
        else:
            root = find_archive_root(archive)

    price_xlsx = root / 'جدول اسعار الخدمات لسنة -2026 new.xlsx'
    catalog = load_price_catalog(price_xlsx)

    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(price_xlsx, SOURCE_DIR / price_xlsx.name)
    svc_xlsx = root / 'جدول  طلبات خدمات  المشتركين.xlsx'
    if svc_xlsx.exists():
        shutil.copy2(svc_xlsx, SOURCE_DIR / svc_xlsx.name)

    with open(SERVICES_JSON, encoding='utf-8') as f:
        data = json.load(f)

    # حفظ كتالوج الأسعار
    price_out = {
        'year': 2026,
        'currency': 'IQD',
        'source': f'data/source/{price_xlsx.name}',
        'catalog': list(catalog.values()),
    }
    with open(ROOT / 'data' / 'service_prices.json', 'w', encoding='utf-8') as f:
        json.dump(price_out, f, ensure_ascii=False, indent=2)

    updated_pricing = flowcharts = 0
    for code, spec in SERVICE_PRICE_MAP.items():
        if code not in data.get('services', {}):
            continue
        svc = data['services'][code]
        svc['pricing'] = build_pricing(code, spec, catalog)
        updated_pricing += 1

        # لا نستبدل guide.procedure تلقائياً — الاستخراج من docx قد يُنتج نصاً ناقصاً
        # ويُفضَّل الإبقاء على الإجراءات المُدقَّقة في services.json

    flowcharts = sync_flowcharts(root, data['services'])

    with open(SERVICES_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')

    print(f'✓ pricing محدّث لـ {updated_pricing} خدمة')
    print(f'✓ flowcharts منسوخة: {flowcharts}')
    print(f'✓ data/service_prices.json + data/source/')


if __name__ == '__main__':
    main()
